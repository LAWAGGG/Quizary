from collections import Counter
from io import BytesIO
from datetime import datetime
import re

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.dependencies import get_current_user, verify_form_owner
from app.models.form import Form
from app.models.submission import Submission, SubmissionStatus
from app.models.answer import Answer
from app.models.answer_option import AnswerOption
from app.models.question import Question, QuestionType
from app.models.question_option import QuestionOption
from app.models.user import User
from app.services.grading import grade_answer
from app.services.session_expiry import auto_submit_expired_for_form
from app.utils import to_naive_utc, fmt_dt, now_wib, _delete_file
from app.services.grading import grade_submission
from app.schemas.results import (
    ResultDeleteRequest,
    ResultStatusRequest,
    ResultBulkStatusRequest,
    ResultItem,
    ResultListResponse,
    AnalyticsResponse,
    PerQuestionStat,
    ScoreDistribution,
    OptionChoice,
    QuestionStat,
    QuestionHighlight,
    DashboardResponse,
    RecentForm,
    SubmissionTrend,
)

router = APIRouter(tags=["results & dashboard"])


@router.get("/forms/{form_id}/results", response_model=ResultListResponse)
def list_results(
    form: Form = Depends(verify_form_owner),
    status_filter: str | None = Query(None, alias="status"),
    sort: str | None = Query(None, alias="sort"),
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    db: Session = Depends(get_db),
):
    # Sweep sesi yang kedaluwarsa supaya tidak ada submission yang menganggur
    # selamanya di status `in_progress` — dikonversi jadi `auto_submitted`.
    auto_submit_expired_for_form(db, form)
    q = db.query(Submission).filter(Submission.form_id == form.id)
    if status_filter:
        if status_filter not in SubmissionStatus.__members__:
            raise HTTPException(status_code=422, detail="status must be in_progress, submitted, auto_submitted, or cheating")
        q = q.filter(Submission.status == SubmissionStatus[status_filter])

    if sort == "score_desc":
        q = q.order_by(Submission.score.desc(), Submission.submitted_at.asc(), Submission.id.asc())
    elif sort == "score_asc":
        q = q.order_by(Submission.score.asc())
    else:
        q = q.order_by(Submission.created_at.desc())

    total = q.count()
    subs = q.offset((page - 1) * per_page).limit(per_page).all()

    # Rank global (urutan score DESC) selalu dihitung, tidak tergantung sort —
    # supaya kolom Rank tampil di seluruh mode sorting. Konsisten dengan leaderboard.
    rank_map: dict[int, int] = {}
    if form.type.value == "quiz":
        ordered = (
            db.query(Submission.id)
            .filter(
                Submission.form_id == form.id,
                Submission.status.in_([SubmissionStatus.submitted, SubmissionStatus.auto_submitted, SubmissionStatus.cheating]),
            )
            .order_by(Submission.score.desc(), Submission.submitted_at.asc(), Submission.id.asc())
            .all()
        )
        rank_map = {sid: i + 1 for i, (sid,) in enumerate(ordered)}

    answer_summary: dict[int, str] = {}
    if form.type.value != "quiz" and subs:
        q_ids = [row[0] for row in db.query(Question.id).filter(Question.form_id == form.id, Question.is_deleted.is_(False)).all()]
        opt_text = {o.id: _strip_html(o.option_text) for o in db.query(QuestionOption).filter(QuestionOption.question_id.in_(q_ids)).all()}
        answers = db.query(Answer).filter(
            Answer.submission_id.in_([s.id for s in subs]),
            Answer.question_id.in_(q_ids),
        ).all()
        ans_opt = db.query(AnswerOption).filter(AnswerOption.answer_id.in_([a.id for a in answers])).all()
        ao_by_a: dict[int, list[str]] = {}
        for ao in ans_opt:
            text = opt_text.get(ao.option_id)
            if text:
                ao_by_a.setdefault(ao.answer_id, []).append(text)
        q_type = {row.id: row.type for row in db.query(Question.id, Question.type).filter(Question.id.in_(q_ids)).all()}
        by_sub: dict[int, list[str]] = {}
        for a in answers:
            if q_type.get(a.question_id) in (QuestionType.multiple_choice, QuestionType.checkbox, QuestionType.dropdown):
                text = ", ".join(ao_by_a.get(a.id, []))
            elif q_type.get(a.question_id) == QuestionType.file_upload:
                text = a.answer_file or ""
            else:
                text = a.answer_text or ""
            if text:
                by_sub.setdefault(a.submission_id, []).append(text)
        for s in subs:
            joined = " · ".join(by_sub.get(s.id, [])[:3])
            answer_summary[s.id] = joined[:100]

    return ResultListResponse(
        data=[ResultItem(
            submission_id=s.id,
            respondent_name=s.respondent_name,
            is_creator=s.user_id == form.user_id,
            score=float(s.score) if s.score is not None else None,
            max_score=float(s.max_score) if s.max_score is not None else None,
            status=s.status.value,
            submitted_at=fmt_dt(to_naive_utc(s.submitted_at)),
            answer_summary=answer_summary.get(s.id, ""),
            rank=rank_map.get(s.id),
            cheat_reason=s.cheat_reason,
        ) for i, s in enumerate(subs)],
        meta={"total": total, "page": page, "per_page": per_page},
    )


# ── DELETE /forms/{form_id}/results ──────────────────────────────────────────

@router.delete("/forms/{form_id}/results")
def delete_results(
    body: ResultDeleteRequest,
    form: Form = Depends(verify_form_owner),
    db: Session = Depends(get_db),
):
    """Bulk-delete hasil. Id yang bukan milik form ini / sudah terhapus
    diabaikan — respons berisi jumlah yang benar-benar terhapus."""
    subs = (
        db.query(Submission)
        .filter(Submission.form_id == form.id, Submission.id.in_(body.submission_ids))
        .all()
    )
    for s in subs:
        # File jawaban upload ikut dibersihkan dari disk (baris DB hilang via cascade).
        for (path,) in (
            db.query(Answer.answer_file)
            .filter(Answer.submission_id == s.id, Answer.answer_file.isnot(None))
            .all()
        ):
            _delete_file(path)
        db.delete(s)
    db.commit()
    return {"deleted": len(subs), "message": f"{len(subs)} hasil berhasil dihapus"}


# ── PATCH /forms/{form_id}/results/{submission_id}/status ────────────────────

@router.patch("/forms/{form_id}/results/{submission_id}/status")
def set_result_status(
    body: ResultStatusRequest,
    submission_id: int,
    form: Form = Depends(verify_form_owner),
    db: Session = Depends(get_db),
):
    """Creator mengatur ulang status hasil secara universal — membuka submission
    yang terkunci/ditandai salah, mensahkan, atau memvonis curang (nilai 0)."""
    sub = db.query(Submission).filter(
        Submission.id == submission_id, Submission.form_id == form.id
    ).first()
    if not sub:
        raise HTTPException(status_code=404, detail="Hasil tidak ditemukan")

    now = now_wib()
    if body.status == "in_progress":
        sub.status = SubmissionStatus.in_progress
        sub.started_at = now
        sub.submitted_at = None
        sub.score = None
        message = "Submission dibuka kembali — responden dapat melanjutkan"
    elif body.status == "submitted":
        sub.status = SubmissionStatus.submitted
        sub.submitted_at = sub.submitted_at or now
        grade_submission(db, sub, form)
        message = "Submission disahkan"
    else:  # cheating
        sub.status = SubmissionStatus.cheating
        sub.submitted_at = sub.submitted_at or now
        grade_submission(db, sub, form)
        sub.score = 0
        message = "Submission dinilai curang (nilai 0)"

    sub.updated_at = now
    db.commit()
    return {
        "submission_id": sub.id,
        "status": sub.status.value,
        "score": float(sub.score) if sub.score is not None else None,
        "message": message,
    }


@router.patch("/forms/{form_id}/results/status")
def set_bulk_result_status(
    body: ResultBulkStatusRequest,
    form: Form = Depends(verify_form_owner),
    db: Session = Depends(get_db),
):
    """Bulk update status submission."""
    subs = (
        db.query(Submission)
        .filter(Submission.form_id == form.id, Submission.id.in_(body.submission_ids))
        .all()
    )
    if not subs:
        return {"updated": 0, "message": "Tidak ada hasil yang diubah"}

    now = now_wib()
    for sub in subs:
        if body.status == "in_progress":
            sub.status = SubmissionStatus.in_progress
            sub.started_at = now
            sub.submitted_at = None
            sub.score = None
        elif body.status == "submitted":
            sub.status = SubmissionStatus.submitted
            sub.submitted_at = sub.submitted_at or now
            grade_submission(db, sub, form)
        else:  # cheating
            sub.status = SubmissionStatus.cheating
            sub.submitted_at = sub.submitted_at or now
            grade_submission(db, sub, form)
            sub.score = 0
        
        sub.updated_at = now

    db.commit()
    return {"updated": len(subs), "message": f"{len(subs)} hasil berhasil diperbarui ke {body.status}"}


@router.get("/forms/{form_id}/analytics", response_model=AnalyticsResponse)
def get_analytics(form: Form = Depends(verify_form_owner), db: Session = Depends(get_db)):
    subs = db.query(Submission).filter(
        Submission.form_id == form.id,
        Submission.status.in_([SubmissionStatus.submitted, SubmissionStatus.auto_submitted, SubmissionStatus.cheating]),
    ).all()

    total = len(subs)
    questions = db.query(Question).filter(Question.form_id == form.id, Question.is_deleted.is_(False)).order_by(Question.order_index).all()
    sub_ids = [s.id for s in subs]

    if total == 0:
        return AnalyticsResponse(type=form.type.value, total_participants=0)

    all_answers = db.query(Answer).filter(
        Answer.question_id.in_([q.id for q in questions]),
        Answer.submission_id.in_(sub_ids),
    ).all()

    answer_by_q: dict[int, list[Answer]] = {}
    for a in all_answers:
        answer_by_q.setdefault(a.question_id, []).append(a)

    # ── Form type: answer-frequency analysis ────────────────────────────────
    if form.type.value != "quiz":
        question_stats: list[QuestionStat] = []
        total_answers = 0
        for q in questions:
            answers = answer_by_q.get(q.id, [])
            # Jawaban file (tipe file_upload) disimpan di answer_file, bukan
            # answer_text/selected_options — wajib dihitung juga, kalau tidak
            # soal upload selalu tampil "0 answered" di analytics.
            non_empty = [
                a for a in answers
                if (a.answer_text and a.answer_text.strip()) or a.selected_options or a.answer_file
            ]
            answered = len(non_empty)
            total_answers += answered

            opts = sorted(q.options, key=lambda o: o.order_index or 0)
            opt_counts: Counter = Counter()
            for a in non_empty:
                for ao in a.selected_options:
                    opt_counts[ao.option_id] += 1

            breakdown = [
                OptionChoice(
                    option_id=o.id,
                    option_text=o.option_text,
                    count=opt_counts.get(o.id, 0),
                    pct=round(opt_counts.get(o.id, 0) / answered * 100, 1) if answered else 0,
                )
                for o in opts
            ]
            most = max(breakdown, key=lambda b: b.count) if breakdown and any(b.count for b in breakdown) else None

            sample_answers = [
                (a.answer_text.strip() if a.answer_text and a.answer_text.strip()
                 else f"[File] {a.answer_file.rsplit('/', 1)[-1]}")
                for a in non_empty
                if (a.answer_text and a.answer_text.strip()) or a.answer_file
            ][:5]

            question_stats.append(QuestionStat(
                question_id=q.id,
                question_text=q.question_text,
                type=q.type.value,
                answered=answered,
                skipped=total - answered,
                most_selected=most.option_text if most else None,
                most_selected_count=most.count if most else 0,
                most_selected_pct=most.pct if most else 0,
                option_breakdown=breakdown,
                sample_answers=sample_answers,
            ))

        completion = total_answers / (total * len(questions)) if total and questions else 0
        return AnalyticsResponse(
            type="form",
            total_participants=total,
            total_answers=total_answers,
            completion_rate=round(completion, 2),
            avg_answers=round(total_answers / total, 1) if total else 0,
            question_stats=question_stats,
        )

    # ── Quiz type: score & correctness analysis ─────────────────────────────
    scores = [float(s.score or 0) for s in subs]
    avg = sum(scores) / total

    per_q = []
    total_correct = total_answers = 0
    for q in questions:
        answers = answer_by_q.get(q.id, [])
        correct = wrong = 0
        for a in answers:
            # Live-grade at read time (grade_answer) so stored is_correct is
            # only a hint: historical/ungraded rows still count instead of 0.
            verdict, _ = grade_answer(a, q)
            if verdict is True:
                correct += 1
            elif verdict is False:
                wrong += 1
        per_q.append(PerQuestionStat(
            question_id=q.id,
            question_text=q.question_text,
            correct_count=correct,
            wrong_count=wrong,
        ))
        total_correct += correct
        total_answers += correct + wrong

    rate = total_correct / total_answers if total_answers else 0

    sorted_scores = sorted(scores)
    n = len(sorted_scores)
    median = (sorted_scores[n // 2] + sorted_scores[(n - 1) // 2]) / 2 if n else 0
    above_avg = sum(1 for s in scores if s > avg) / total if total else 0

    # Quiz pool = 100 (auto) atau dinormalisasi ke 100 (manual).
    # Bucket berbasis persentase agar relevan untuk kedua mode.
    dist: dict[str, int] = {}
    for s in scores:
        if s <= 25:
            key = "0-25"
        elif s <= 50:
            key = "26-50"
        elif s <= 75:
            key = "51-75"
        else:
            key = "76-100"
        dist[key] = dist.get(key, 0) + 1

    # ── Pace (Duration) calculation ─────────────────────────────────────────
    durations = []
    for s in subs:
        if s.started_at and s.submitted_at:
            diff = (s.submitted_at - s.started_at).total_seconds()
            if diff >= 0:
                durations.append(diff)

    avg_duration = int(sum(durations) / len(durations)) if durations else None
    fastest_duration = int(min(durations)) if durations else None

    # ── Item Difficulty Diagnostics (Easiest & Hardest) ───────────────────────
    evaluated_questions = []
    for idx, q_stat in enumerate(per_q):
        attempt_total = q_stat.correct_count + q_stat.wrong_count
        if attempt_total > 0:
            acc = q_stat.correct_count / attempt_total
            evaluated_questions.append({
                "order_index": idx + 1,
                "question_text": q_stat.question_text,
                "accuracy": round(acc * 100, 1),
            })

    easiest_q = None
    hardest_q = None
    if evaluated_questions:
        sorted_by_acc = sorted(evaluated_questions, key=lambda x: x["accuracy"])
        hardest_q = QuestionHighlight(
            order_index=sorted_by_acc[0]["order_index"],
            question_text=sorted_by_acc[0]["question_text"],
            accuracy=sorted_by_acc[0]["accuracy"],
        )
        easiest_q = QuestionHighlight(
            order_index=sorted_by_acc[-1]["order_index"],
            question_text=sorted_by_acc[-1]["question_text"],
            accuracy=sorted_by_acc[-1]["accuracy"],
        )

    return AnalyticsResponse(
        type="quiz",
        total_participants=total,
        average_score=round(avg, 2),
        median_score=round(median, 2),
        highest_score=max(scores),
        lowest_score=min(scores),
        above_average_pct=round(above_avg, 2),
        correct_rate=round(rate, 2),
        wrong_rate=round(1 - rate, 2),
        avg_duration_seconds=avg_duration,
        fastest_duration_seconds=fastest_duration,
        easiest_question=easiest_q,
        hardest_question=hardest_q,
        score_distribution=[ScoreDistribution(range=k, count=v) for k, v in sorted(dist.items())],
        per_question_stats=per_q,
    )


def _strip_html(text) -> str:
    """Buang tag HTML dari teks rich (question/option) untuk export yang bersih."""
    import re
    text = str(text or "")
    text = re.sub(r"<[^>]*>", "", text)
    return text.strip()


# Karakter pembuka yang membuat Excel menafsirkan sel sebagai formula.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value):
    """Netralkan Excel formula injection dari konten milik responden.

    openpyxl menulis string berawalan '=' sebagai formula aktif; jawaban seperti
    '=WEBSERVICE(...)' akan dieksekusi saat file dibuka. Prefiks apostrof
    memaksa sel dirender sebagai teks biasa (perilaku standar Excel).
    """
    if isinstance(value, str) and value[:1] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _export_columns(form: Form, subs: list[Submission], db: Session, base_url: str | None = None):
    """Build dynamic export: one column per question + Dikirim/Skor/Status."""
    questions = db.query(Question).filter(Question.form_id == form.id, Question.is_deleted.is_(False)).order_by(Question.order_index).all()
    q_ids = [q.id for q in questions]
    headers = [_safe_cell(_strip_html(q.question_text) or f"Soal {i+1}") for i, q in enumerate(questions)] + ["Dikirim", "Skor", "Status"]

    if not questions:
        return questions, headers, []

    opt_text = {o.id: _strip_html(o.option_text) for o in db.query(QuestionOption).filter(QuestionOption.question_id.in_(q_ids)).all()}
    answers = db.query(Answer).filter(
        Answer.question_id.in_(q_ids),
        Answer.submission_id.in_([s.id for s in subs]),
    ).all()

    ans_opt = db.query(AnswerOption).filter(AnswerOption.answer_id.in_([a.id for a in answers])).all()
    ao_by_answer: dict[int, list[str]] = {}
    for ao in ans_opt:
        text = opt_text.get(ao.option_id)
        if text:
            ao_by_answer.setdefault(ao.answer_id, []).append(text)

    def _file_link(path: str | None) -> str:
        """Path relatif → URL penuh agar bisa diklik di Excel."""
        if not path:
            return ""
        if path.startswith("http://") or path.startswith("https://"):
            return path
        if base_url:
            return f"{base_url.rstrip('/')}/uploads/{path.lstrip('/')}"
        return path

    q_by_id = {q.id: q for q in questions}
    answer_map: dict[tuple[int, int], str] = {}
    for a in answers:
        q = q_by_id.get(a.question_id)
        if not q:
            continue
        if q.type in (QuestionType.multiple_choice, QuestionType.checkbox, QuestionType.dropdown):
            answer_map[(a.submission_id, a.question_id)] = ", ".join(ao_by_answer.get(a.id, []))
        elif q.type == QuestionType.file_upload:
            answer_map[(a.submission_id, a.question_id)] = _file_link(a.answer_file) or _strip_html(a.answer_text) or ""
        else:
            answer_map[(a.submission_id, a.question_id)] = _strip_html(a.answer_text) or ""

    rows = []
    for s in subs:
        row = [_safe_cell(answer_map.get((s.id, q.id), "") or "-") for q in questions]
        row.append(fmt_dt(to_naive_utc(s.submitted_at)) or "-")
        row.append(float(s.score) if s.score is not None else "-")
        row.append(s.status.value)
        rows.append(row)
    return questions, headers, rows


@router.get("/forms/{form_id}/export/excel")
def export_excel(
    request: Request,
    form: Form = Depends(verify_form_owner),
    db: Session = Depends(get_db),
    status: str | None = Query(None, alias="status"),
    sort: str | None = Query(None, alias="sort"),
):
    # Sesi kedaluwarsa dikonversi dulu jadi auto_submitted supaya datanya ikut
    # terekspor. Submission in_progress yang masih aktif tetap turut diekspor
    # (skor kosong) agar pengerjaan yang belum selesai ikut terlihat.
    auto_submit_expired_for_form(db, form)
    q = db.query(Submission).filter(Submission.form_id == form.id)

    # Filter by status — same as GET /results, includes locked when asked
    if status:
        if status not in SubmissionStatus.__members__:
            raise HTTPException(status_code=422, detail="status must be in_progress, submitted, auto_submitted, cheating or locked")
        q = q.filter(Submission.status == SubmissionStatus[status])
    else:
        q = q.filter(Submission.status.in_([
            SubmissionStatus.in_progress,
            SubmissionStatus.submitted,
            SubmissionStatus.auto_submitted,
            SubmissionStatus.cheating,
            SubmissionStatus.locked,
        ]))

    # Sort — mirror list_results so export matches what user sees
    if sort == "score_desc":
        q = q.order_by(Submission.score.desc(), Submission.submitted_at.asc(), Submission.id.asc())
    elif sort == "score_asc":
        q = q.order_by(Submission.score.asc(), Submission.submitted_at.asc(), Submission.id.asc())
    elif sort == "newest":
        q = q.order_by(Submission.created_at.desc())
    elif sort == "oldest":
        q = q.order_by(Submission.created_at.asc())
    elif sort == "status":
        q = q.order_by(Submission.status.asc(), Submission.created_at.desc())
    elif sort:
        raise HTTPException(status_code=422, detail="sort must be score_desc, score_asc, newest, oldest or status")
    else:
        q = q.order_by(Submission.created_at.desc())

    subs = q.all()

    base_url = str(request.base_url).rstrip("/")
    questions, headers, rows = _export_columns(form, subs, db, base_url)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Soal file_upload → hyperlink klikable (display = nama file, target = URL
    # penuh). Excel render biru-underline; user klik langsung buka file.
    file_col_idx = {i for i, qq in enumerate(questions) if qq.type == QuestionType.file_upload}
    link_font = Font(color="0563C1", underline="single")
    for row in ws.iter_rows(min_row=2, max_row=1 + len(rows)):
        for i in file_col_idx:
            cell = row[i]
            url = cell.value
            if isinstance(url, str) and url.startswith(("http://", "https://")):
                fname = url.rsplit("/", 1)[-1] or "Lihat File"
                cell.value = fname
                cell.hyperlink = url
                cell.font = link_font

    # Styling: header di-highlight (fill primary, teks putih tebal) + border tipis
    # di semua sel + wrap text, supaya tabel terbaca jelas walau teks panjang.
    header_fill = PatternFill("solid", fgColor="6C5CE7")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = wrap
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = wrap
    ws.freeze_panes = "A2"

    # Auto width per kolom: ikuti konten terpanjang (header/isi), dibatasi 50
    # karakter supaya tabel tak melar; teks panjang wrap ke bawah dalam kolom lebar.
    for i, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = now_wib().strftime("%Y-%m-%d")
    safe_title = re.sub(r"[^\w\-]+", "_", _strip_html(form.title)).strip("_") or form.short_code
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_title}_{today}.xlsx"},
    )


@router.get("/dashboard/summary", response_model=DashboardResponse)
def dashboard_summary(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    forms = db.query(Form).filter(Form.user_id == user.id).all()
    form_ids = [f.id for f in forms]

    subs = db.query(Submission).filter(Submission.form_id.in_(form_ids)).all() if form_ids else []

    # Count unique respondents by user_id or ip_address
    respondent_keys: set = set()
    for s in subs:
        respondent_keys.add(f"u:{s.user_id}" if s.user_id else f"ip:{s.ip_address}")

    recent = sorted(forms, key=lambda f: f.created_at or datetime(2000, 1, 1), reverse=True)[:5]
    sub_count_by_form = Counter(s.form_id for s in subs)
    recent_data = [
        RecentForm(id=f.id, title=_strip_html(f.title), status=f.status.value, submission_count=sub_count_by_form.get(f.id, 0))
        for f in recent
    ]

    # Submission trend per form (bukan per hari) — tiap bar menampilkan satu
    # form, sehingga di chart langsung terlihat form mana yang paling banyak
    # mendapat jawaban. Dibatasi 10 teratas agar tidak penuh.
    form_count = Counter(s.form_id for s in subs)
    title_by_id = {f.id: _strip_html(f.title) for f in forms}
    trend = [
        SubmissionTrend(form_id=fid, title=title_by_id.get(fid, "(deleted)"), count=c)
        for fid, c in sorted(form_count.items(), key=lambda x: -x[1])[:10]
    ]
    return DashboardResponse(
        total_forms=len(forms),
        total_quiz=sum(1 for f in forms if f.type.value == "quiz"),
        total_submissions=len(subs),
        total_respondents=len(respondent_keys),
        recent_forms=recent_data,
        submission_trend=trend,
    )
